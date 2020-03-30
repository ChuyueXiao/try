import openpyxl
# 写入的代码:
wb = openpyxl.Workbook()
sheet = wb.active   # wb.active就是获取这个工作薄的活动表，通常就是第一个工作表。
sheet.title = 'marvel'
sheet['A1'] = 'universal'
rows = [['captain america','great hunk','iron man'],['black widow','captain marvel','spider man']]
for i in rows:
    sheet.append(i)
print(rows)
wb.save('Marvel World.xlsx')
# 读取的代码：
wb = openpyxl.load_workbook('Marvel World.xlsx')  # 调用openpyxl.load_workbook()函数，打开文件
sheet = wb['marvel']  # 获取“Marvel World.xlsx”工作薄中名为“marvel”的工作表
# sheetnames是用来获取工作薄所有工作表的名字的。如果不知道工作薄到底有几个工作表，就可以把工作表的名字都打印出来
sheetname = wb.sheetnames
print(sheetname)
A1_cell = sheet['A1']
A1_value = A1_cell.value
print(A1_value)

# 引用csv模块
import csv
#  写入:
# 调用open()函数打开csv文件，传入参数：文件名“demo.csv”、写入模式“w”、newline=''、encoding='utf-8'。
csv_file = open('demo.csv','w',newline='',encoding='utf-8')
writer = csv.writer(csv_file)  # 用csv.writer()函数创建一个writer对象
# 调用writer对象的writerow()方法，可以在csv文件里一行行写入文字：
writer.writerow(['电影','豆瓣评分'])
writer.writerow(['银河护卫队','8.0'])
writer.writerow(['复仇者联盟','8.1'])
csv_file.close()  # 写入完成后，关闭文件就可以了

# 读取：
csv_file = open('demo.csv','r',newline='',encoding='utf-8')
reader=csv.reader(csv_file)  # 用csv.reader()函数创建一个reader对象。
for row in reader:
    print(row)